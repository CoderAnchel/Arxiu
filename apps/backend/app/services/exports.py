"""XLSX/CSV export builders.

Each builder receives an `AsyncSession` plus an entity id and returns the raw
bytes of the workbook/CSV. Sheets follow a consistent design:

- `Resum`        — one-row overview (codi, nom, dates, totals)
- `Matrícules`   — alumne-level rows, one per matrícula
- `Mòduls`       — mòdul-level structure
- `Notes RA`     — long-format: matr × ra × avaluació, with nota + comentari
- `Notes mòdul`  — final notes (auto/manual)

The exporters are deliberately verbose: they do their own SELECTs rather than
sharing the response schemas with the read endpoints, because the export
format diverges (long format vs nested JSON) and we want it to be stable
even if response shapes change.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.exceptions import NotFound
from app.models.audit_log import AuditLog
from app.models.catalog import Cicle, CursAcademic, Modul, Ra
from app.models.grading import Avaluacio, QualificacioModul, QualificacioRa
from app.models.people import (
    Alumne,
    AssignacioDocent,
    GrupClasse,
    Matricula,
    TutorLegal,
)
from app.models.user import User

# ----- Styles -------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="E7E5E0")
HEADER_FONT = Font(bold=True, size=10, name="Inter", color="3A3733")
TITLE_FONT = Font(bold=True, size=14, name="Inter", color="2A2724")
SUBTITLE_FONT = Font(italic=True, size=11, name="Inter", color="6C6862")
MONO_FONT = Font(name="JetBrains Mono", size=9, color="6C6862")
SUSP_FONT = Font(bold=True, color="B5453E")
EXC_FONT = Font(bold=True, color="2F7A41")


def _setup_workbook() -> Workbook:
    wb = Workbook()
    # remove the default sheet — we'll add our own with explicit names
    wb.remove(wb.active)
    return wb


def _header_row(ws: Worksheet, row: int, headers: list[str]) -> None:
    for col, label in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=label)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22


def _autosize(ws: Worksheet, min_width: int = 8, max_width: int = 50) -> None:
    """Crude auto-size: scan column values, pick longest str, clamp."""
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        longest = min_width
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            if len(s) > longest:
                longest = min(len(s), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = longest + 2


def _wb_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _format_nota(n: Decimal | float | None) -> float | str:
    if n is None:
        return ""
    return float(n)


# ============================================================================
# Alumne expedient — one sheet per matrícula × mòdul, plus summary
# ============================================================================


async def export_alumne(session: AsyncSession, alumne_id: int) -> tuple[bytes, str]:
    alumne = await session.get(Alumne, alumne_id)
    if alumne is None or alumne.deleted_at is not None:
        raise NotFound("alumne_not_found")

    tutors = list(
        (
            await session.execute(
                select(TutorLegal).where(TutorLegal.alumne_id == alumne.id)
            )
        )
        .scalars()
        .all()
    )

    matricules = list(
        (
            await session.execute(
                select(Matricula)
                .where(Matricula.alumne_id == alumne.id, Matricula.deleted_at.is_(None))
                .order_by(Matricula.curs_acad_id.desc())
            )
        )
        .scalars()
        .all()
    )

    wb = _setup_workbook()

    # --- Sheet 1: Resum -----------------------------------------------------
    ws = wb.create_sheet("Resum")
    ws["A1"] = f"Expedient acadèmic · {alumne.cognoms}, {alumne.nom}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    ws["A2"] = f"Generat el {datetime.now():%Y-%m-%d %H:%M}"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:F2")

    rows = [
        ("DNI", alumne.dni or ""),
        ("RALC/NIA", alumne.ralc),
        ("Email", alumne.email or ""),
        ("Telèfon", alumne.telefon or ""),
        ("Data de naixement", alumne.data_naixement.isoformat() if alumne.data_naixement else ""),
        ("Tutors legals", "; ".join(f"{t.nom} <{t.email or '—'}>" for t in tutors) or "—"),
        ("Matrícules totals", len(matricules)),
    ]
    for i, (k, v) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=k).font = HEADER_FONT
        ws.cell(row=i, column=2, value=v)
    _autosize(ws)

    # --- Per-matrícula sheets ----------------------------------------------
    for matr in matricules:
        curs = await session.get(CursAcademic, matr.curs_acad_id)
        cicle = await session.get(Cicle, matr.cicle_id)
        grup = await session.get(GrupClasse, matr.grup_id)
        if curs is None or cicle is None or grup is None:
            continue

        avals = list(
            (
                await session.execute(
                    select(Avaluacio)
                    .where(
                        Avaluacio.curs_acad_id == curs.id,
                        Avaluacio.deleted_at.is_(None),
                    )
                    .order_by(Avaluacio.ordre)
                )
            )
            .scalars()
            .all()
        )

        moduls = list(
            (
                await session.execute(
                    select(Modul)
                    .where(
                        Modul.cicle_id == cicle.id,
                        Modul.curs == matr.curs,
                        Modul.deleted_at.is_(None),
                    )
                    .options(selectinload(Modul.ras))
                    .order_by(Modul.codi)
                )
            )
            .scalars()
            .unique()
            .all()
        )

        qras = list(
            (
                await session.execute(
                    select(QualificacioRa).where(QualificacioRa.matricula_id == matr.id)
                )
            )
            .scalars()
            .all()
        )
        ra_lookup = {(q.avaluacio_id, q.ra_id): q for q in qras}

        qmoduls = list(
            (
                await session.execute(
                    select(QualificacioModul).where(
                        QualificacioModul.matricula_id == matr.id
                    )
                )
            )
            .scalars()
            .all()
        )
        modul_lookup = {(q.avaluacio_id, q.modul_id): q for q in qmoduls}

        sheet_name = f"{curs.nom}_{grup.codi}"[:31]  # Excel limit
        wsm = wb.create_sheet(sheet_name)
        wsm["A1"] = f"{curs.nom} · {grup.codi} · {cicle.codi} {cicle.nom} · {matr.curs}r curs"
        wsm["A1"].font = TITLE_FONT
        wsm.merge_cells(f"A1:{get_column_letter(4 + len(avals))}1")

        headers = ["Mòdul", "Codi RA", "Descripció", "Pes %"] + [a.nom for a in avals] + ["Mitjana", "Final manual"]
        _header_row(wsm, 3, headers)

        r = 4
        for mod in moduls:
            ras_sorted = sorted(mod.ras, key=lambda x: x.ordre)
            for ra in ras_sorted:
                wsm.cell(row=r, column=1, value=mod.codi + " · " + mod.nom)
                wsm.cell(row=r, column=2, value=ra.codi)
                wsm.cell(row=r, column=3, value=ra.descripcio)
                wsm.cell(row=r, column=4, value=float(ra.pes) if ra.pes is not None else 0.0)
                for i, aval in enumerate(avals, start=5):
                    q = ra_lookup.get((aval.id, ra.id))
                    cell = wsm.cell(row=r, column=i, value=_format_nota(q.nota if q else None))
                    if q and q.nota is not None:
                        if q.nota < 5:
                            cell.font = SUSP_FONT
                        elif q.nota >= 9:
                            cell.font = EXC_FONT
                r += 1
            # mòdul-level mean + manual override row (per avaluació)
            wsm.cell(row=r, column=1, value=f"{mod.codi} · NOTA MÒDUL").font = HEADER_FONT
            wsm.cell(row=r, column=4, value="(ponderada)")
            for i, aval in enumerate(avals, start=5):
                qm = modul_lookup.get((aval.id, mod.id))
                ras_sorted = sorted(mod.ras, key=lambda x: x.ordre)
                weighted: list[tuple[float, float]] = []
                for ra in ras_sorted:
                    q = ra_lookup.get((aval.id, ra.id))
                    if q is not None and q.nota is not None and ra.pes is not None:
                        weighted.append((float(q.nota), float(ra.pes)))
                if weighted:
                    total_w = sum(p for _, p in weighted) or 1.0
                    mean = sum(n * p for n, p in weighted) / total_w
                    val = round(mean, 2)
                else:
                    val = ""
                cell = wsm.cell(row=r, column=i, value=val)
                cell.font = Font(bold=True)
                if qm is not None and qm.nota is not None:
                    # manual override — show with a marker prefix
                    cell.value = f"{float(qm.nota):.2f} (manual)"
                    cell.font = Font(bold=True, italic=True, color="2F7A41")
            r += 1
            r += 1  # blank line between mòduls
        _autosize(wsm)

    filename = f"expedient_{alumne.cognoms.replace(' ', '_')}_{alumne.nom.replace(' ', '_')}.xlsx"
    return _wb_to_bytes(wb), filename


# ============================================================================
# Grup expedient — all students × all moduls
# ============================================================================


async def export_grup(session: AsyncSession, grup_id: int) -> tuple[bytes, str]:
    grup = await session.get(GrupClasse, grup_id)
    if grup is None or grup.deleted_at is not None:
        raise NotFound("grup_not_found")

    curs = await session.get(CursAcademic, grup.curs_acad_id)
    cicle = await session.get(Cicle, grup.cicle_id)
    tutor = await session.get(User, grup.tutor_user_id) if grup.tutor_user_id else None

    matricules = list(
        (
            await session.execute(
                select(Matricula)
                .where(
                    Matricula.grup_id == grup.id,
                    Matricula.deleted_at.is_(None),
                )
                .order_by(Matricula.id)
            )
        )
        .scalars()
        .all()
    )
    alumnes = []
    for m in matricules:
        a = await session.get(Alumne, m.alumne_id)
        if a:
            alumnes.append((m, a))
    alumnes.sort(key=lambda x: (x[1].cognoms, x[1].nom))

    moduls = list(
        (
            await session.execute(
                select(Modul)
                .where(
                    Modul.cicle_id == grup.cicle_id,
                    Modul.curs == grup.curs,
                    Modul.deleted_at.is_(None),
                )
                .options(selectinload(Modul.ras))
                .order_by(Modul.codi)
            )
        )
        .scalars()
        .unique()
        .all()
    )

    avals = list(
        (
            await session.execute(
                select(Avaluacio)
                .where(
                    Avaluacio.curs_acad_id == grup.curs_acad_id,
                    Avaluacio.deleted_at.is_(None),
                )
                .order_by(Avaluacio.ordre)
            )
        )
        .scalars()
        .all()
    )

    matr_ids = [m.id for m, _ in alumnes]

    qra_lookup: dict[tuple[int, int, int], QualificacioRa] = {}
    qmod_lookup: dict[tuple[int, int, int], QualificacioModul] = {}
    if matr_ids:
        qras = list(
            (
                await session.execute(
                    select(QualificacioRa).where(QualificacioRa.matricula_id.in_(matr_ids))
                )
            )
            .scalars()
            .all()
        )
        for q in qras:
            qra_lookup[(q.matricula_id, q.avaluacio_id, q.ra_id)] = q
        qmoduls = list(
            (
                await session.execute(
                    select(QualificacioModul).where(QualificacioModul.matricula_id.in_(matr_ids))
                )
            )
            .scalars()
            .all()
        )
        for q in qmoduls:
            qmod_lookup[(q.matricula_id, q.avaluacio_id, q.modul_id)] = q

    wb = _setup_workbook()

    # --- Resum ---
    ws = wb.create_sheet("Resum")
    ws["A1"] = f"Grup {grup.codi} · {cicle.codi if cicle else '—'} · {curs.nom if curs else '—'}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = f"Generat el {datetime.now():%Y-%m-%d %H:%M}"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:F2")
    rows = [
        ("Cicle", f"{cicle.codi} · {cicle.nom}" if cicle else "—"),
        ("Curs (1r/2n)", grup.curs),
        ("Curs acadèmic", curs.nom if curs else "—"),
        ("Tutor/a", f"{tutor.nom} {tutor.cognoms}" if tutor else "—"),
        ("Alumnes matriculats", len(alumnes)),
        ("Mòduls", len(moduls)),
        ("Avaluacions", len(avals)),
    ]
    for i, (k, v) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=k).font = HEADER_FONT
        ws.cell(row=i, column=2, value=v)
    _autosize(ws)

    # --- Alumnes ---
    wsa = wb.create_sheet("Alumnes")
    _header_row(wsa, 1, ["Cognoms", "Nom", "DNI", "RALC", "Email", "Tipus", "Estat"])
    for i, (m, a) in enumerate(alumnes, start=2):
        wsa.cell(row=i, column=1, value=a.cognoms)
        wsa.cell(row=i, column=2, value=a.nom)
        wsa.cell(row=i, column=3, value=a.dni or "")
        wsa.cell(row=i, column=4, value=a.ralc)
        wsa.cell(row=i, column=5, value=a.email or "")
        wsa.cell(row=i, column=6, value=m.tipus.value if hasattr(m.tipus, "value") else str(m.tipus))
        wsa.cell(row=i, column=7, value=m.estat.value if hasattr(m.estat, "value") else str(m.estat))
    _autosize(wsa)

    # --- One sheet per mòdul: alumnes × (RA × avaluacions) + final ---
    for mod in moduls:
        ras_sorted = sorted(mod.ras, key=lambda x: x.ordre)
        sheet_name = f"{mod.codi}"[:31]
        wsm = wb.create_sheet(sheet_name)
        wsm["A1"] = f"{mod.codi} · {mod.nom} ({mod.hores} h · {len(ras_sorted)} RA)"
        wsm["A1"].font = TITLE_FONT
        wsm.merge_cells(f"A1:{get_column_letter(2 + len(avals) * len(ras_sorted))}1")

        # Header rows: aval | RA1 RA2 ... | FINAL
        headers: list[str] = ["Cognoms", "Nom"]
        for aval in avals:
            for ra in ras_sorted:
                headers.append(f"{aval.nom} · {ra.codi}")
            headers.append(f"{aval.nom} · FINAL")
        _header_row(wsm, 3, headers)

        for i, (m, a) in enumerate(alumnes, start=4):
            wsm.cell(row=i, column=1, value=a.cognoms)
            wsm.cell(row=i, column=2, value=a.nom)
            col = 3
            for aval in avals:
                for ra in ras_sorted:
                    q = qra_lookup.get((m.id, aval.id, ra.id))
                    cell = wsm.cell(row=i, column=col, value=_format_nota(q.nota if q else None))
                    if q and q.nota is not None:
                        if q.nota < 5:
                            cell.font = SUSP_FONT
                        elif q.nota >= 9:
                            cell.font = EXC_FONT
                    col += 1
                # FINAL: manual override else weighted mean
                qm = qmod_lookup.get((m.id, aval.id, mod.id))
                if qm and qm.nota is not None:
                    cell = wsm.cell(row=i, column=col, value=float(qm.nota))
                    cell.font = Font(bold=True, italic=True, color="2F7A41")
                else:
                    weighted: list[tuple[float, float]] = []
                    for ra in ras_sorted:
                        q = qra_lookup.get((m.id, aval.id, ra.id))
                        if q is not None and q.nota is not None and ra.pes is not None:
                            weighted.append((float(q.nota), float(ra.pes)))
                    if weighted:
                        tw = sum(p for _, p in weighted) or 1.0
                        cell = wsm.cell(row=i, column=col, value=round(sum(n * p for n, p in weighted) / tw, 2))
                        cell.font = Font(bold=True)
                col += 1
        _autosize(wsm)

    filename = f"grup_{grup.codi}_{curs.nom if curs else 'curs'}.xlsx"
    return _wb_to_bytes(wb), filename


# ============================================================================
# Grup × Mòdul — single spreadsheet matching the QualifsPage view
# ============================================================================


async def export_grup_modul(
    session: AsyncSession, grup_id: int, modul_id: int, avaluacio_id: int | None = None
) -> tuple[bytes, str]:
    grup = await session.get(GrupClasse, grup_id)
    modul = await session.get(Modul, modul_id)
    if grup is None or modul is None:
        raise NotFound("grup_or_modul_not_found")
    if grup.curs != modul.curs:
        raise NotFound("curs_mismatch")

    matricules = list(
        (
            await session.execute(
                select(Matricula)
                .where(
                    Matricula.grup_id == grup.id,
                    Matricula.curs == modul.curs,
                    Matricula.deleted_at.is_(None),
                )
            )
        )
        .scalars()
        .all()
    )
    alumnes: list[tuple[Matricula, Alumne]] = []
    for m in matricules:
        a = await session.get(Alumne, m.alumne_id)
        if a:
            alumnes.append((m, a))
    alumnes.sort(key=lambda x: (x[1].cognoms, x[1].nom))

    ras = list(
        (
            await session.execute(
                select(Ra)
                .where(Ra.modul_id == modul.id, Ra.deleted_at.is_(None))
                .order_by(Ra.ordre)
            )
        )
        .scalars()
        .all()
    )

    aval_stmt = select(Avaluacio).where(
        Avaluacio.curs_acad_id == grup.curs_acad_id,
        Avaluacio.deleted_at.is_(None),
    )
    if avaluacio_id is not None:
        aval_stmt = aval_stmt.where(Avaluacio.id == avaluacio_id)
    avals = list((await session.execute(aval_stmt.order_by(Avaluacio.ordre))).scalars().all())

    matr_ids = [m.id for m, _ in alumnes]
    qra_lookup: dict[tuple[int, int, int], QualificacioRa] = {}
    qmod_lookup: dict[tuple[int, int], QualificacioModul] = {}
    if matr_ids:
        qras = list(
            (
                await session.execute(
                    select(QualificacioRa).where(
                        QualificacioRa.matricula_id.in_(matr_ids),
                        QualificacioRa.ra_id.in_([r.id for r in ras]),
                    )
                )
            )
            .scalars()
            .all()
        )
        for q in qras:
            qra_lookup[(q.matricula_id, q.avaluacio_id, q.ra_id)] = q
        qmoduls = list(
            (
                await session.execute(
                    select(QualificacioModul).where(
                        QualificacioModul.matricula_id.in_(matr_ids),
                        QualificacioModul.modul_id == modul.id,
                    )
                )
            )
            .scalars()
            .all()
        )
        for q in qmoduls:
            qmod_lookup[(q.matricula_id, q.avaluacio_id)] = q

    wb = _setup_workbook()
    for aval in avals:
        sheet_name = f"{aval.nom}"[:31]
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = f"{grup.codi} · {modul.codi} {modul.nom} · {aval.nom} ({aval.estat.value if hasattr(aval.estat, 'value') else aval.estat})"
        ws["A1"].font = TITLE_FONT
        ws.merge_cells(f"A1:{get_column_letter(3 + len(ras))}1")

        headers = ["Cognoms", "Nom"] + [f"{r.codi} ({r.pes}%)" for r in ras] + ["FINAL"]
        _header_row(ws, 3, headers)

        for i, (m, a) in enumerate(alumnes, start=4):
            ws.cell(row=i, column=1, value=a.cognoms)
            ws.cell(row=i, column=2, value=a.nom)
            for j, r in enumerate(ras, start=3):
                q = qra_lookup.get((m.id, aval.id, r.id))
                cell = ws.cell(row=i, column=j, value=_format_nota(q.nota if q else None))
                if q and q.nota is not None:
                    if q.nota < 5:
                        cell.font = SUSP_FONT
                    elif q.nota >= 9:
                        cell.font = EXC_FONT
            qm = qmod_lookup.get((m.id, aval.id))
            final_col = 3 + len(ras)
            if qm and qm.nota is not None:
                cell = ws.cell(row=i, column=final_col, value=float(qm.nota))
                cell.font = Font(bold=True, italic=True, color="2F7A41")
            else:
                weighted = [
                    (float(qra_lookup[(m.id, aval.id, r.id)].nota), float(r.pes))
                    for r in ras
                    if (m.id, aval.id, r.id) in qra_lookup
                    and qra_lookup[(m.id, aval.id, r.id)].nota is not None
                    and r.pes is not None
                ]
                if weighted:
                    tw = sum(p for _, p in weighted) or 1.0
                    cell = ws.cell(
                        row=i,
                        column=final_col,
                        value=round(sum(n * p for n, p in weighted) / tw, 2),
                    )
                    cell.font = Font(bold=True)
        _autosize(ws)

    filename = f"qualifs_{grup.codi}_{modul.codi}.xlsx"
    return _wb_to_bytes(wb), filename


# ============================================================================
# Curs acadèmic — summary across all grups
# ============================================================================


async def export_curs(session: AsyncSession, curs_id: int) -> tuple[bytes, str]:
    curs = await session.get(CursAcademic, curs_id)
    if curs is None:
        raise NotFound("curs_not_found")

    grups = list(
        (
            await session.execute(
                select(GrupClasse)
                .where(
                    GrupClasse.curs_acad_id == curs.id,
                    GrupClasse.deleted_at.is_(None),
                )
                .order_by(GrupClasse.codi)
            )
        )
        .scalars()
        .all()
    )

    wb = _setup_workbook()
    ws = wb.create_sheet("Resum")
    ws["A1"] = f"Curs acadèmic {curs.nom}{' (actiu)' if curs.actiu else ''}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    ws["A2"] = f"Generat el {datetime.now():%Y-%m-%d %H:%M}"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:E2")

    _header_row(ws, 4, ["Grup", "Cicle", "Curs", "Tutor/a", "Alumnes"])
    r = 5
    total_alumnes = 0
    for g in grups:
        cicle = await session.get(Cicle, g.cicle_id)
        tutor = await session.get(User, g.tutor_user_id) if g.tutor_user_id else None
        n_alumnes = (
            await session.execute(
                select(Matricula).where(
                    Matricula.grup_id == g.id, Matricula.deleted_at.is_(None)
                )
            )
        ).scalars().all()
        count = len(n_alumnes)
        total_alumnes += count
        ws.cell(row=r, column=1, value=g.codi)
        ws.cell(row=r, column=2, value=f"{cicle.codi}" if cicle else "")
        ws.cell(row=r, column=3, value=f"{g.curs}r")
        ws.cell(row=r, column=4, value=f"{tutor.nom} {tutor.cognoms}" if tutor else "—")
        ws.cell(row=r, column=5, value=count)
        r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = HEADER_FONT
    ws.cell(row=r, column=5, value=total_alumnes).font = HEADER_FONT
    _autosize(ws)

    filename = f"curs_{curs.nom}.xlsx"
    return _wb_to_bytes(wb), filename


# ============================================================================
# Cicle — structure + alumnes who have ever taken it
# ============================================================================


async def export_cicle(session: AsyncSession, cicle_id: int) -> tuple[bytes, str]:
    cicle = await session.get(Cicle, cicle_id)
    if cicle is None or cicle.deleted_at is not None:
        raise NotFound("cicle_not_found")

    moduls = list(
        (
            await session.execute(
                select(Modul)
                .where(Modul.cicle_id == cicle.id, Modul.deleted_at.is_(None))
                .options(selectinload(Modul.ras))
                .order_by(Modul.curs, Modul.codi)
            )
        )
        .scalars()
        .unique()
        .all()
    )

    wb = _setup_workbook()
    ws = wb.create_sheet("Estructura")
    ws["A1"] = f"Cicle {cicle.codi} · {cicle.nom}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    ws["A2"] = f"Nivell {cicle.nivell.value if hasattr(cicle.nivell, 'value') else cicle.nivell} · durada {cicle.durada} curs/os"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:E2")

    _header_row(ws, 4, ["Curs", "Mòdul", "Nom", "Hores", "RAs"])
    r = 5
    for m in moduls:
        ws.cell(row=r, column=1, value=f"{m.curs}r")
        ws.cell(row=r, column=2, value=m.codi)
        ws.cell(row=r, column=3, value=m.nom)
        ws.cell(row=r, column=4, value=m.hores)
        ws.cell(row=r, column=5, value=len(m.ras))
        r += 1
        for ra in sorted(m.ras, key=lambda x: x.ordre):
            ws.cell(row=r, column=2, value=ra.codi).font = MONO_FONT
            ws.cell(row=r, column=3, value=ra.descripcio)
            ws.cell(row=r, column=4, value=f"{ra.pes}%")
            r += 1
        r += 1
    _autosize(ws)

    # Alumnes who have ever been matriculated in this cicle
    matricules = list(
        (
            await session.execute(
                select(Matricula).where(
                    Matricula.cicle_id == cicle.id, Matricula.deleted_at.is_(None)
                )
            )
        )
        .scalars()
        .all()
    )
    wsa = wb.create_sheet("Alumnes (històric)")
    _header_row(wsa, 1, ["Cognoms", "Nom", "DNI", "RALC", "Curs", "Curs acadèmic", "Grup", "Estat"])
    rows: list[tuple[str, str, str, str, int, str, str, str]] = []
    for m in matricules:
        a = await session.get(Alumne, m.alumne_id)
        c = await session.get(CursAcademic, m.curs_acad_id)
        g = await session.get(GrupClasse, m.grup_id)
        if a is None:
            continue
        rows.append(
            (
                a.cognoms,
                a.nom,
                a.dni or "",
                a.ralc,
                m.curs,
                c.nom if c else "",
                g.codi if g else "",
                m.estat.value if hasattr(m.estat, "value") else str(m.estat),
            )
        )
    rows.sort(key=lambda x: (x[5], x[6], x[0], x[1]))
    for i, row in enumerate(rows, start=2):
        for j, v in enumerate(row, start=1):
            wsa.cell(row=i, column=j, value=v)
    _autosize(wsa)

    filename = f"cicle_{cicle.codi}.xlsx"
    return _wb_to_bytes(wb), filename


# ============================================================================
# Docent — assignacions + tutoritzacions
# ============================================================================


async def export_docent(session: AsyncSession, user_id: int) -> tuple[bytes, str]:
    user = await session.get(User, user_id)
    if user is None or not user.active:
        raise NotFound("docent_not_found")

    assigs = list(
        (
            await session.execute(
                select(AssignacioDocent).where(
                    AssignacioDocent.user_id == user.id,
                    AssignacioDocent.deleted_at.is_(None),
                )
            )
        )
        .scalars()
        .all()
    )
    tutorships = list(
        (
            await session.execute(
                select(GrupClasse).where(
                    GrupClasse.tutor_user_id == user.id,
                    GrupClasse.deleted_at.is_(None),
                )
            )
        )
        .scalars()
        .all()
    )

    wb = _setup_workbook()
    ws = wb.create_sheet("Resum")
    ws["A1"] = f"{user.nom} {user.cognoms} · {user.role.value if hasattr(user.role, 'value') else user.role}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    rows = [
        ("DNI", user.dni),
        ("Email", user.email),
        ("Departament", user.departament or "—"),
        ("Assignacions", len(assigs)),
        ("Tutoritzacions", len(tutorships)),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = HEADER_FONT
        ws.cell(row=i, column=2, value=v)
    _autosize(ws)

    wsa = wb.create_sheet("Assignacions")
    _header_row(wsa, 1, ["Curs acadèmic", "Grup", "Mòdul codi", "Mòdul nom"])
    for i, a in enumerate(assigs, start=2):
        c = await session.get(CursAcademic, a.curs_acad_id)
        g = await session.get(GrupClasse, a.grup_id)
        m = await session.get(Modul, a.modul_id)
        wsa.cell(row=i, column=1, value=c.nom if c else "")
        wsa.cell(row=i, column=2, value=g.codi if g else "")
        wsa.cell(row=i, column=3, value=m.codi if m else "")
        wsa.cell(row=i, column=4, value=m.nom if m else "")
    _autosize(wsa)

    wst = wb.create_sheet("Tutoritzacions")
    _header_row(wst, 1, ["Curs acadèmic", "Grup", "Cicle"])
    for i, g in enumerate(tutorships, start=2):
        c = await session.get(CursAcademic, g.curs_acad_id)
        cicle = await session.get(Cicle, g.cicle_id)
        wst.cell(row=i, column=1, value=c.nom if c else "")
        wst.cell(row=i, column=2, value=g.codi)
        wst.cell(row=i, column=3, value=f"{cicle.codi} · {cicle.nom}" if cicle else "")
    _autosize(wst)

    safe_name = f"{user.cognoms}_{user.nom}".replace(" ", "_")
    filename = f"docent_{safe_name}.xlsx"
    return _wb_to_bytes(wb), filename


# ============================================================================
# Audit log — CSV (better for tail-style logs)
# ============================================================================


async def export_audit_csv(
    session: AsyncSession,
    *,
    user_id: int | None = None,
    entity: str | None = None,
    action: str | None = None,
    limit: int = 5000,
) -> tuple[bytes, str]:
    stmt = select(AuditLog).order_by(AuditLog.created_at.desc()).limit(limit)
    if user_id is not None:
        stmt = stmt.where(AuditLog.user_id == user_id)
    if entity is not None:
        stmt = stmt.where(AuditLog.entity == entity)
    if action is not None:
        stmt = stmt.where(AuditLog.action == action)
    rows = list((await session.execute(stmt)).scalars().all())

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(
        ["created_at", "user_id", "action", "entity", "entity_id", "ip", "before", "after"]
    )
    for r in rows:
        writer.writerow(
            [
                r.created_at.isoformat() if r.created_at else "",
                r.user_id or "",
                r.action,
                r.entity,
                r.entity_id or "",
                r.ip or "",
                str(r.before_state) if r.before_state else "",
                str(r.after_state) if r.after_state else "",
            ]
        )
    data = ("﻿" + buf.getvalue()).encode("utf-8")
    filename = f"audit_{datetime.now():%Y%m%d_%H%M}.csv"
    return data, filename
