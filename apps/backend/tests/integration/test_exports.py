"""Tests for /api/v1/export/* endpoints.

We assert two things per export:
1. Status + content-type + filename headers are correct
2. The blob is a parsable XLSX (or CSV) — no schema validation of cell
   contents because the structure is verified by services in unit tests.
"""
from __future__ import annotations

import io

import openpyxl
import pytest
from httpx import AsyncClient

from tests.factories import make_admin, make_user
from app.models.user import UserRole


pytestmark = pytest.mark.asyncio


async def _admin_token(client: AsyncClient, db) -> str:
    await make_admin(db, dni="00000000T", password="Admin-Pwd-1!")
    r = await client.post(
        "/api/v1/auth/login",
        json={"identifier": "00000000T", "password": "Admin-Pwd-1!"},
    )
    return r.json()["access_token"]


async def _setup_minimal(client: AsyncClient, h: dict) -> dict:
    curs = (
        await client.post(
            "/api/v1/cursos-academics",
            headers=h,
            json={"nom": "2025-2026", "actiu": True},
        )
    ).json()
    cicle = (
        await client.post(
            "/api/v1/cicles",
            headers=h,
            json={"codi": "DAM", "nom": "DAM", "nivell": "superior", "durada": 2},
        )
    ).json()
    modul = (
        await client.post(
            "/api/v1/moduls",
            headers=h,
            json={
                "cicle_id": cicle["id"],
                "codi": "M03",
                "nom": "Programació",
                "curs": 1,
                "hores": 264,
            },
        )
    ).json()
    ra = (
        await client.post(
            "/api/v1/ras",
            headers=h,
            json={
                "modul_id": modul["id"],
                "ordre": 1,
                "codi": "RA1",
                "descripcio": "Identifica…",
                "pes": 100,
            },
        )
    ).json()
    grup = (
        await client.post(
            "/api/v1/grups",
            headers=h,
            json={
                "codi": "DAM1A",
                "curs_acad_id": curs["id"],
                "cicle_id": cicle["id"],
                "curs": 1,
            },
        )
    ).json()
    alumne = (
        await client.post(
            "/api/v1/alumnes",
            headers=h,
            json={
                "ralc": "R0000001",
                "dni": "11111111H",
                "nom": "Aleix",
                "cognoms": "Vilanova",
            },
        )
    ).json()
    matr = (
        await client.post(
            "/api/v1/matricules",
            headers=h,
            json={
                "alumne_id": alumne["id"],
                "grup_id": grup["id"],
                "cicle_id": cicle["id"],
                "curs": 1,
                "curs_acad_id": curs["id"],
            },
        )
    ).json()
    return {
        "curs": curs,
        "cicle": cicle,
        "modul": modul,
        "ra": ra,
        "grup": grup,
        "alumne": alumne,
        "matricula": matr,
    }


def _is_xlsx(content: bytes) -> bool:
    try:
        openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        return True
    except Exception:
        return False


async def test_export_alumne_xlsx(client: AsyncClient, db) -> None:
    token = await _admin_token(client, db)
    h = {"Authorization": f"Bearer {token}"}
    ctx = await _setup_minimal(client, h)

    r = await client.get(f"/api/v1/export/alumne/{ctx['alumne']['id']}.xlsx", headers=h)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert "attachment" in r.headers["content-disposition"]
    assert _is_xlsx(r.content)


async def test_export_grup_xlsx(client: AsyncClient, db) -> None:
    token = await _admin_token(client, db)
    h = {"Authorization": f"Bearer {token}"}
    ctx = await _setup_minimal(client, h)

    r = await client.get(f"/api/v1/export/grup/{ctx['grup']['id']}.xlsx", headers=h)
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
    # Should at least contain Resum + Alumnes + one mòdul sheet
    sheets = wb.sheetnames
    assert "Resum" in sheets
    assert "Alumnes" in sheets
    assert any(s == "M03" for s in sheets)


async def test_export_grup_modul_xlsx(client: AsyncClient, db) -> None:
    token = await _admin_token(client, db)
    h = {"Authorization": f"Bearer {token}"}
    ctx = await _setup_minimal(client, h)
    # Need an avaluació for the per-aval sheet
    await client.post(
        "/api/v1/avaluacions",
        headers=h,
        json={"curs_acad_id": ctx["curs"]["id"], "nom": "1a Avaluació", "ordre": 1},
    )

    r = await client.get(
        f"/api/v1/export/grup/{ctx['grup']['id']}/modul/{ctx['modul']['id']}.xlsx",
        headers=h,
    )
    assert r.status_code == 200
    assert _is_xlsx(r.content)


async def test_export_grup_modul_rejects_curs_mismatch(client: AsyncClient, db) -> None:
    token = await _admin_token(client, db)
    h = {"Authorization": f"Bearer {token}"}
    ctx = await _setup_minimal(client, h)
    # Create a 2n-curs mòdul of the same cicle
    m2 = (
        await client.post(
            "/api/v1/moduls",
            headers=h,
            json={
                "cicle_id": ctx["cicle"]["id"],
                "codi": "M06",
                "nom": "Accés a dades",
                "curs": 2,
                "hores": 165,
            },
        )
    ).json()
    r = await client.get(
        f"/api/v1/export/grup/{ctx['grup']['id']}/modul/{m2['id']}.xlsx", headers=h
    )
    # Service raises NotFound("curs_mismatch") → 404
    assert r.status_code == 404


async def test_export_curs_xlsx(client: AsyncClient, db) -> None:
    token = await _admin_token(client, db)
    h = {"Authorization": f"Bearer {token}"}
    ctx = await _setup_minimal(client, h)
    r = await client.get(f"/api/v1/export/curs/{ctx['curs']['id']}.xlsx", headers=h)
    assert r.status_code == 200
    assert _is_xlsx(r.content)


async def test_export_cicle_xlsx(client: AsyncClient, db) -> None:
    token = await _admin_token(client, db)
    h = {"Authorization": f"Bearer {token}"}
    ctx = await _setup_minimal(client, h)
    r = await client.get(f"/api/v1/export/cicle/{ctx['cicle']['id']}.xlsx", headers=h)
    assert r.status_code == 200
    assert _is_xlsx(r.content)


async def test_export_docent_xlsx_admin(client: AsyncClient, db) -> None:
    token = await _admin_token(client, db)
    h = {"Authorization": f"Bearer {token}"}
    prof = await make_user(
        db, dni="11111111A", email="p@inslaferreria.cat", role=UserRole.PROFESSOR
    )
    r = await client.get(f"/api/v1/export/docent/{prof.id}.xlsx", headers=h)
    assert r.status_code == 200
    assert _is_xlsx(r.content)


async def test_export_docent_xlsx_professor_only_own(client: AsyncClient, db) -> None:
    other_admin = await make_admin(db, dni="00000000T", password="Admin-Pwd-1!")
    prof = await make_user(
        db,
        dni="11111111A",
        email="p@inslaferreria.cat",
        role=UserRole.PROFESSOR,
        password="Prof-Pwd-1!",
    )
    r = await client.post(
        "/api/v1/auth/login",
        json={"identifier": "11111111A", "password": "Prof-Pwd-1!"},
    )
    prof_token = r.json()["access_token"]
    h = {"Authorization": f"Bearer {prof_token}"}

    # Own → 200
    own = await client.get(f"/api/v1/export/docent/{prof.id}.xlsx", headers=h)
    assert own.status_code == 200

    # Other (admin) → 403
    nope = await client.get(f"/api/v1/export/docent/{other_admin.id}.xlsx", headers=h)
    assert nope.status_code == 403


async def test_export_audit_csv(client: AsyncClient, db) -> None:
    token = await _admin_token(client, db)
    h = {"Authorization": f"Bearer {token}"}
    # generate some auditable events
    await client.post(
        "/api/v1/cursos-academics", headers=h, json={"nom": "2024-2025"}
    )
    r = await client.get("/api/v1/export/audit.csv", headers=h)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("text/csv")
    # CSV header line present
    assert b"created_at,user_id,action" in r.content


async def test_export_curs_requires_admin(client: AsyncClient, db) -> None:
    await make_admin(db, dni="00000000T", password="Admin-Pwd-1!")
    await make_user(
        db,
        dni="11111111A",
        email="p@inslaferreria.cat",
        role=UserRole.PROFESSOR,
        password="Prof-Pwd-1!",
    )
    r = await client.post(
        "/api/v1/auth/login",
        json={"identifier": "11111111A", "password": "Prof-Pwd-1!"},
    )
    prof_h = {"Authorization": f"Bearer {r.json()['access_token']}"}
    # Setup as admin would, then try as prof — use 1 because the cicle id can't matter
    r2 = await client.get("/api/v1/export/curs/1.xlsx", headers=prof_h)
    assert r2.status_code == 403
